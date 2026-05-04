from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional
import os
import urllib.error
import urllib.request


from core.config.constants import HTTP_DOWNLOAD_MAX_BYTES, HTTP_READ_TIMEOUT_SEC
from core.paths import FONTS_DIR, LEGACY_ASSETS_FONTS, REPO_ROOT

from ..reporting.report_model import build_usage_report
from ..reporting.state import DashboardRangeState
from ..tui.log_console import console

GOOGLE_INTER_FILENAME = "Inter_18pt-Regular.ttf"
GOOGLE_INTER_URL = "https://raw.githubusercontent.com/google/fonts/main/ofl/inter/Inter%5Bopsz,wght%5D.ttf"


def _candidate_fonts(project_root: Optional[Path] = None) -> list[Path]:
    candidates: list[Path] = []
    if project_root is not None:
        candidates.append(Path(project_root) / GOOGLE_INTER_FILENAME)
        candidates.append(Path(project_root) / "fonts" / GOOGLE_INTER_FILENAME)
    windir = os.environ.get("WINDIR") or r"C:\Windows"
    candidates.append(Path(windir) / "Fonts" / GOOGLE_INTER_FILENAME)
    candidates.append(FONTS_DIR / GOOGLE_INTER_FILENAME)
    candidates.append(LEGACY_ASSETS_FONTS / GOOGLE_INTER_FILENAME)
    candidates.append(REPO_ROOT / "fonts" / GOOGLE_INTER_FILENAME)
    candidates.append(Path(__file__).resolve().parents[1] / GOOGLE_INTER_FILENAME)
    candidates.append(Path(__file__).resolve().parents[1] / "fonts" / GOOGLE_INTER_FILENAME)
    seen: set[str] = set()
    out: list[Path] = []
    for p in candidates:
        k = str(p).lower()
        if k not in seen:
            seen.add(k)
            out.append(p)
    return out


def _download_google_font(target_dir: Path) -> Optional[Path]:
    target_dir.mkdir(parents=True, exist_ok=True)
    out = target_dir / GOOGLE_INTER_FILENAME
    if out.exists() and out.stat().st_size > 0:
        return out
    tmp = out.with_suffix(out.suffix + ".part")
    try:
        req = urllib.request.Request(
            GOOGLE_INTER_URL, headers={"User-Agent": "aiteam/6.2"}
        )
        with urllib.request.urlopen(req, timeout=HTTP_READ_TIMEOUT_SEC) as resp:  # nosec B310
            written = 0
            with open(tmp, "wb") as fh:
                while True:
                    chunk = resp.read(64 * 1024)
                    if not chunk:
                        break
                    written += len(chunk)
                    if written > HTTP_DOWNLOAD_MAX_BYTES:
                        raise ValueError(
                            f"font exceeds cap {HTTP_DOWNLOAD_MAX_BYTES}"
                        )
                    fh.write(chunk)
        tmp.replace(out)
        if out.exists() and out.stat().st_size > 0:
            return out
    except (OSError, urllib.error.URLError, ValueError):
        try:
            if tmp.exists():
                tmp.unlink()
        except OSError:
            pass
        return None
    return None


def _select_font(project_root: Optional[Path] = None) -> tuple[Optional[Path], str]:
    for p in _candidate_fonts(project_root):
        if p.exists() and p.is_file():
            source = "workspace" if project_root and str(p).startswith(str(project_root)) else "windows"
            return p, source
    downloaded = _download_google_font(Path(project_root or Path.cwd()) / "fonts")
    if downloaded:
        return downloaded, "google-download"
    return None, "txt"


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                value = str(cell.value) if cell.value is not None else ""
            except (TypeError, ValueError, AttributeError):
                value = ""
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 48)


def _style_sheet(ws) -> None:
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7C9D6")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    _autofit(ws)


def export_excel(project_root: Path, range_state: DashboardRangeState) -> Path:
    import openpyxl
    from openpyxl.styles import Font

    report = build_usage_report(range_state)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = project_root / f"ai_team_usage_{ts}.xlsx"
    wb = openpyxl.Workbook()

    ws_kpi = wb.active
    ws_kpi.title = "KPI"
    ws_kpi.merge_cells("A1:B1")
    ws_kpi["A1"] = "AI Team — Usage (dashboard export)"
    ws_kpi["A1"].font = Font(bold=True, size=14)
    ws_kpi["A2"] = "Range label"
    ws_kpi["B2"] = report.label
    ws_kpi["A3"] = "From"
    ws_kpi["B3"] = report.since.isoformat() if report.since else ""
    ws_kpi["A4"] = "To"
    ws_kpi["B4"] = report.until.isoformat() if report.until else ""
    ws_kpi["A5"] = "Generated"
    ws_kpi["B5"] = report.generated_at.isoformat(timespec="seconds")
    ws_kpi["A7"] = "Total requests"
    ws_kpi["B7"] = report.total_requests
    ws_kpi["A8"] = "Total tokens"
    ws_kpi["B8"] = report.total_tokens
    ws_kpi["A9"] = "Total spend (USD)"
    ws_kpi["B9"] = round(report.total_spend, 6)
    ws_kpi["A10"] = "Input tokens"
    ws_kpi["B10"] = report.prompt_tokens
    ws_kpi["A11"] = "Output tokens"
    ws_kpi["B11"] = report.completion_tokens
    ws_kpi["A12"] = "CLI turns"
    ws_kpi["B12"] = report.cli_turns
    for addr in ("B7", "B8", "B10", "B11", "B12"):
        ws_kpi[addr].number_format = "#,##0"
    ws_kpi["B9"].number_format = "#,##0.000000"
    ws_kpi.column_dimensions["A"].width = 22
    ws_kpi.column_dimensions["B"].width = 44

    ws = wb.create_sheet("Summary")
    ws.append(["Range", range_state.label])
    ws.append(["From", range_state.since.isoformat() if range_state.since else ""])
    ws.append(["To", range_state.until.isoformat() if range_state.until else ""])
    ws.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    ws.append([])
    ws.append(["Batch #", "Timestamp", "Mode", "Tokens", "Spend (USD)", "Requests"])
    for b in report.batches:
        tot = b.get("totals") or {}
        ws.append(
            [
                b.get("batch_idx"),
                b.get("timestamp"),
                b.get("mode"),
                int(tot.get("total_tokens", 0)),
                float(b.get("cost_usd", 0.0)),
                len(b.get("usage_rows") or []),
            ]
        )
    for row in range(7, 7 + len(report.batches)):
        for col in (4, 5, 6):
            ws.cell(row=row, column=col).number_format = "#,##0.000000" if col == 5 else "#,##0"

    ws2 = wb.create_sheet("RoleModel")
    ws2.append(["Role", "Model", "Requests", "Tokens", "Spend (USD)"])
    for row in report.by_role_model:
        ws2.append([row.role, row.model, row.requests, row.tokens, row.cost_usd])
    for r in range(2, 2 + len(report.by_role_model)):
        for c in (3, 4, 5):
            ws2.cell(row=r, column=c).number_format = "#,##0.000000" if c == 5 else "#,##0"

    ws3 = wb.create_sheet("RawRows")
    ws3.append(["Timestamp", "Role", "Model", "Prompt", "Completion", "Total", "Spend (USD)"])
    for r in report.raw_rows:
        ws3.append(
            [
                r.get("timestamp"),
                str(r.get("role_key") or r.get("agent") or "unknown"),
                str(r.get("model") or ""),
                int(r.get("prompt_tokens", 0) or 0),
                int(r.get("completion_tokens", 0) or 0),
                int(r.get("total_tokens", 0) or 0),
                float(r.get("cost_usd", 0.0) or 0.0),
            ]
        )
    for r in range(2, min(2 + len(report.raw_rows), 5002)):
        for c in (4, 5, 6, 7):
            ws3.cell(row=r, column=c).number_format = "#,##0.000000" if c == 7 else "#,##0"

    ws4 = wb.create_sheet("PeriodSummary")
    ws4.append(["Period", "Requests", "Tokens", "Spend (USD)"])
    for period_name, stats in sorted(report.period_summary.items()):
        ws4.append(
            [
                period_name,
                int(stats.get("requests", 0)),
                int(stats.get("tokens", 0)),
                float(stats.get("spend", 0.0)),
            ]
        )
    for r in range(2, 2 + len(report.period_summary)):
        for c in (2, 3, 4):
            ws4.cell(row=r, column=c).number_format = "#,##0.000000" if c == 4 else "#,##0"

    for sheet in (ws, ws2, ws3, ws4):
        _style_sheet(sheet)

    wb.save(out)
    console.print(f"[green]XLSX → {out}[/green]")
    return out

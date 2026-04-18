from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional
import os
import urllib.error
import urllib.request
from collections import defaultdict

from utils import tracker

from .state import DashboardRangeState
from .utils import default_range
from .render import console

GOOGLE_INTER_FILENAME = "Inter_18pt-Regular.ttf"
GOOGLE_INTER_BOLD_FILENAME = "Inter_24pt-Bold.ttf"
GOOGLE_INTER_URL = "https://raw.githubusercontent.com/google/fonts/main/ofl/inter/Inter%5Bopsz,wght%5D.ttf"


def _candidate_fonts(project_root: Optional[Path] = None) -> list[Path]:
    candidates: list[Path] = []
    if project_root is not None:
        candidates.append(Path(project_root) / GOOGLE_INTER_FILENAME)
        candidates.append(Path(project_root) / "fonts" / GOOGLE_INTER_FILENAME)
    windir = os.environ.get("WINDIR") or r"C:\Windows"
    candidates.append(Path(windir) / "Fonts" / GOOGLE_INTER_FILENAME)
    candidates.append(Path(__file__).resolve().parents[1] / GOOGLE_INTER_FILENAME)
    candidates.append(Path(__file__).resolve().parents[1] / "fonts" / GOOGLE_INTER_FILENAME)
    seen: set[str] = set()
    out: list[Path] = []
    for p in candidates:
        k = str(p).lower()
        if k not in seen:
            seen.add(k)
            out.append(p)
    return out


def _download_google_font(target_dir: Path) -> Optional[Path]:
    target_dir.mkdir(parents=True, exist_ok=True)
    out = target_dir / GOOGLE_INTER_FILENAME
    if out.exists() and out.stat().st_size > 0:
        return out
    try:
        urllib.request.urlretrieve(GOOGLE_INTER_URL, out)
        if out.exists() and out.stat().st_size > 0:
            return out
    except (OSError, urllib.error.URLError, ValueError):
        return None
    return None


def _select_font(project_root: Optional[Path] = None) -> tuple[Optional[Path], str]:
    for p in _candidate_fonts(project_root):
        if p.exists() and p.is_file():
            source = "workspace" if project_root and str(p).startswith(str(project_root)) else "windows"
            return p, source
    downloaded = _download_google_font(Path(project_root or Path.cwd()) / "fonts")
    if downloaded:
        return downloaded, "google-download"
    return None, "txt"


def build_xlsx_export_lines(range_state: DashboardRangeState) -> list[dict]:
    since = range_state.since or default_range()[0]
    until = range_state.until or default_range()[1]
    batches = tracker.summarize_tokens_by_cli_batches(since, until)
    rows = []
    for batch in batches:
        rows.append({
            "batch_idx": batch.get("batch_idx"),
            "timestamp": batch.get("timestamp"),
            "mode": batch.get("mode"),
            "total_tokens": int((batch.get("totals") or {}).get("total_tokens", 0)),
            "cost_usd": float(batch.get("cost_usd", 0.0)),
            "requests": len(batch.get("usage_rows") or []),
        })
    return rows


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                value = str(cell.value) if cell.value is not None else ""
            except (TypeError, ValueError, AttributeError):
                value = ""
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 42)


def _style_sheet(ws) -> None:
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7C9D6")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    _autofit(ws)


def build_role_model_breakdown(range_state: DashboardRangeState) -> list[dict]:
    since = range_state.since or default_range()[0]
    until = range_state.until or default_range()[1]
    rows = tracker.read_usage_rows_timerange(since, until)
    agg: dict[tuple[str, str], dict[str, float | int]] = {}
    for r in rows:
        role = str(r.get("role_key") or r.get("agent") or "unknown")
        model = str(r.get("model") or "(unknown)")
        key = (role, model)
        if key not in agg:
            agg[key] = {"role": role, "model": model, "requests": 0, "tokens": 0, "cost_usd": 0.0}
        agg[key]["requests"] += 1
        agg[key]["tokens"] += int(r.get("total_tokens", 0) or 0)
        agg[key]["cost_usd"] += float(r.get("cost_usd", 0.0) or 0.0)
    return sorted(agg.values(), key=lambda x: (-int(x["tokens"]), str(x["role"]), str(x["model"])))


def build_raw_rows(range_state: DashboardRangeState) -> list[dict]:
    since = range_state.since or default_range()[0]
    until = range_state.until or default_range()[1]
    rows = tracker.read_usage_rows_timerange(since, until)
    out: list[dict] = []
    for r in rows:
        out.append({
            "timestamp": r.get("timestamp"),
            "role": str(r.get("role_key") or r.get("agent") or "unknown"),
            "model": str(r.get("model") or ""),
            "prompt_tokens": int(r.get("prompt_tokens", 0) or 0),
            "completion_tokens": int(r.get("completion_tokens", 0) or 0),
            "total_tokens": int(r.get("total_tokens", 0) or 0),
            "cost_usd": float(r.get("cost_usd", 0.0) or 0.0),
        })
    return out


def build_period_summary(range_state: DashboardRangeState) -> dict[str, dict[str, float | int]]:
    return tracker.get_period_usage()


def export_excel(project_root: Path, range_state: DashboardRangeState) -> Path:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for xlsx export") from exc

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = project_root / f"ai_team_usage_{ts}.xlsx"
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Range", range_state.label])
    ws.append(["From", range_state.since.isoformat() if range_state.since else ""])
    ws.append(["To", range_state.until.isoformat() if range_state.until else ""])
    ws.append(["Generated", datetime.now().isoformat(timespec="seconds")])

    ws.append([])
    ws.append(["Batch #", "Timestamp", "Mode", "Tokens", "Spend", "Requests"])
    for row in build_xlsx_export_lines(range_state):
        ws.append([row["batch_idx"], row["timestamp"], row["mode"], row["total_tokens"], row["cost_usd"], row["requests"]])

    ws2 = wb.create_sheet("RoleModel")
    ws2.append(["Role", "Model", "Requests", "Tokens", "Spend"])
    for row in build_role_model_breakdown(range_state):
        ws2.append([row["role"], row["model"], row["requests"], row["tokens"], row["cost_usd"]])

    ws3 = wb.create_sheet("RawRows")
    ws3.append(["Timestamp", "Role", "Model", "Prompt", "Completion", "Total", "Spend"])
    for row in build_raw_rows(range_state):
        ws3.append([row["timestamp"], row["role"], row["model"], row["prompt_tokens"], row["completion_tokens"], row["total_tokens"], row["cost_usd"]])

    ws4 = wb.create_sheet("PeriodSummary")
    summary = build_period_summary(range_state)
    ws4.append(["Period", "Requests", "Tokens", "Spend"])
    for period_name, stats in summary.items():
        ws4.append([period_name, stats.get("requests", 0), stats.get("tokens", 0), stats.get("spend", 0.0)])

    for ws in (ws, ws2, ws3, ws4):
        _style_sheet(ws)

    wb.save(out)
    console.print(f"[green]XLSX → {out}[/green]")
    return out
